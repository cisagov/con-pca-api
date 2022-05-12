"""Safelist utils."""
# Standard Python Libraries
import subprocess  # nosec

# Third-Party Libraries
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

# cisagov Libraries
from utils.templates import get_deception_level


def generate_safelist_file(
    subscription_id,
    phish_header,
    domains,
    ips,
    templates,
    reporting_password,
    simulation_url,
):
    """Generate excel file with safelisting information."""
    # Create workbook
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Set fonts
    header_font = Font(name="Helvetica Neue", bold=True, size=14)
    regular_font = Font(name="Helvetica Neue", size=14)
    top_align = Alignment(vertical="top")
    cola = ws.column_dimensions["A"]
    cola.font = regular_font
    colb = ws.column_dimensions["B"]
    colb.font = regular_font

    # Phish Header
    ws["A1"] = "Cisa-Phish"
    ws["A1"].font = header_font
    ws["B1"] = phish_header
    ws["B1"].font = regular_font

    # Simulation URL
    ws["A2"] = "Simulation URL"
    ws["A2"].font = header_font
    ws["B2"] = simulation_url
    ws["B2"].font = regular_font

    # Sending Domains
    ws["A3"] = "Sending Domains"
    ws["A3"].font = header_font
    ws["A3"].alignment = top_align
    if domains:
        for i, domain in enumerate(domains):
            ws[f"B{3 + i}"] = domain
            ws[f"B{3 + i}"].font = regular_font
        ws.merge_cells(f"A3:A{3 + len(domains) - 1}")

    # Sending Ips
    ip_start = 3 + len(domains)
    ws[f"A{ip_start}"] = "Sending IP Addresses"
    ws[f"A{ip_start}"].alignment = top_align
    ws[f"A{ip_start}"].font = header_font
    if ips:
        for i, ip in enumerate(ips):
            ws[f"B{ip_start + i}"] = ip
            ws[f"B{ip_start + i}"].font = regular_font
        ws.merge_cells(f"A{ip_start}:A{ip_start + len(ips) - 1}")

    # Templates
    template_start = 3 + len(domains) + (1 if not ips else len(ips)) + 1
    ws[f"A{template_start}"] = "Template Subject"
    ws[f"B{template_start}"] = "Deception Level"
    ws[f"A{template_start}"].font = header_font
    ws[f"B{template_start}"].font = header_font
    for i, template in enumerate(templates):
        ws[f"A{template_start + i + 1}"] = template["subject"]
        ws[f"A{template_start + i + 1}"].font = regular_font
        ws[f"B{template_start + i + 1}"] = get_deception_level(
            template["deception_score"]
        ).title()
        ws[f"B{template_start + i + 1}"].font = regular_font

    # Expand columns in workbook
    for cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in cells)
        new_column_letter = get_column_letter(cells[0].column)
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length * 1.5

    filepath = f"/var/www/tmp_safelist_{subscription_id}.xslx"
    wb.save(filepath)

    # Protect excel workbook
    if reporting_password:
        args = [
            "/var/www/msoffice-crypt.exe",
            "-e",
            "-p",
            reporting_password,
            filepath,
            filepath,
        ]
        subprocess.run(args)  # nosec

    return filepath
