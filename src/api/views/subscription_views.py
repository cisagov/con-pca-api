"""Subscription Views."""
# Standard Python Libraries
import logging
import os
import subprocess  # nosec

# Third-Party Libraries
from flask import jsonify, request, send_file
from flask.views import MethodView
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

# cisagov Libraries
from api.manager import (
    CustomerManager,
    CycleManager,
    SubscriptionManager,
    TargetManager,
)
from utils.safelist_testing import test_subscription
from utils.subscriptions import (
    create_subscription_name,
    get_random_phish_header,
    start_subscription,
    stop_subscription,
)
from utils.templates import get_deception_level
from utils.valid import is_subscription_valid

subscription_manager = SubscriptionManager()
customer_manager = CustomerManager()
cycle_manager = CycleManager()
target_manager = TargetManager()


class SubscriptionsView(MethodView):
    """SubscriptionsView."""

    def get(self):
        """Get."""
        parameters = dict(request.args)
        parameters = subscription_manager.get_query(parameters)

        if request.args.get("template"):
            cycles = cycle_manager.all(
                params={"template_ids": request.args["template"]},
                fields=["subscription_id"],
            )
            subscription_ids = list({c["subscription_id"] for c in cycles})
            parameters["$or"] = [
                {"_id": {"$in": subscription_ids}},
                {"templates_selected": request.args["template"]},
            ]

        parameters["archived"] = {"$in": [False, None]}
        if request.args.get("archived", "").lower() == "true":
            parameters["archived"] = True

        return jsonify(
            subscription_manager.all(
                params=parameters,
                fields=[
                    "_id",
                    "customer_id",
                    "name",
                    "status",
                    "start_date",
                    "cycle_length_minutes",
                    "active",
                    "archived",
                    "primary_contact",
                    "admin_email",
                    "created",
                    "created_by",
                    "updated",
                    "updated_by",
                ],
            )
        )

    def post(self):
        """Post."""
        subscription = request.json
        customer = customer_manager.get(document_id=subscription["customer_id"])
        subscription["name"] = create_subscription_name(customer)
        subscription["status"] = "created"
        subscription["phish_header"] = get_random_phish_header()
        response = subscription_manager.save(subscription)
        response["name"] = subscription["name"]
        return jsonify(response)


class SubscriptionView(MethodView):
    """SubscriptionView."""

    def get(self, subscription_id):
        """Get."""
        return jsonify(subscription_manager.get(document_id=subscription_id))

    def put(self, subscription_id):
        """Put."""
        if "target_email_list" in request.json:
            if not request.json["target_email_list"]:
                subscription = subscription_manager.get(
                    document_id=subscription_id, fields=["status"]
                )
                if subscription["status"] in ["queued", "running"]:
                    return (
                        jsonify(
                            {"error": "Subscription started, target list required."}
                        ),
                        400,
                    )
        subscription_manager.update(document_id=subscription_id, data=request.json)
        return jsonify({"success": True})

    def delete(self, subscription_id):
        """Delete."""
        subscription_manager.delete(document_id=subscription_id)
        cycle_manager.delete(params={"subscription_id": subscription_id})
        target_manager.delete(params={"subscription_id": subscription_id})
        return jsonify({"success": True})


class SubscriptionLaunchView(MethodView):
    """SubscriptionLaunchView."""

    def get(self, subscription_id):
        """Launch a subscription."""
        resp, status_code = start_subscription(subscription_id)
        return jsonify(resp), status_code

    def delete(self, subscription_id):
        """Stop a subscription."""
        return jsonify(stop_subscription(subscription_id))


class SubscriptionTestView(MethodView):
    """SubscriptionTestView."""

    def get(self, subscription_id):
        """Get test results for a subscription."""
        return jsonify(
            subscription_manager.get(
                document_id=subscription_id, fields=["test_results"]
            ).get("test_results", [])
        )

    def post(self, subscription_id):
        """Launch a test for the subscription."""
        resp, status_code = test_subscription(subscription_id, request.json["contacts"])
        return jsonify(resp), status_code


class SubscriptionValidView(MethodView):
    """SubscriptionValidView."""

    def post(self):
        """Post."""
        data = request.json
        return jsonify(
            is_subscription_valid(
                data["target_count"],
                data["cycle_minutes"],
            )
        )


class SubscriptionHeaderView(MethodView):
    """SubscriptionHeaderView."""

    def get(self, subscription_id):
        """
        Get.

        Rotate the subscription phishing header.
        """
        new_header = get_random_phish_header()
        subscription_manager.update(
            document_id=subscription_id, data={"phish_header": new_header}
        )
        return jsonify({"phish_header": new_header})


class SubscriptionSafelistExportView(MethodView):
    """SubscriptionSafelistExportView."""

    def post(self, subscription_id):
        """
        Post.

        Get an excel file with safelist attributes in it.
        """
        data = request.json

        phish_header = data["phish_header"]
        domains = data["domains"]
        ips = data["ips"]
        templates = data["templates"]
        reporting_password = data["password"]
        simulation_url = data.get("simulation_url", "")

        # Create workbook
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Set fonts
        header_font = Font(name="Helvetica Neue", bold=True, size=14)
        regular_font = Font(name="Helvetica Neue", size=14)
        top_align = Alignment(vertical="top")
        cola = ws.column_dimensions["A"]
        cola.font = regular_font
        colb = ws.column_dimensions["B"]
        colb.font = regular_font

        # Phish Header
        ws["A1"] = "Cisa-Phish"
        ws["A1"].font = header_font
        ws["B1"] = phish_header
        ws["B1"].font = regular_font

        # Simulation URL
        ws["A2"] = "Simulation URL"
        ws["A2"].font = header_font
        ws["B2"] = simulation_url
        ws["B2"].font = regular_font

        # Sending Domains
        ws["A3"] = "Sending Domains"
        ws["A3"].font = header_font
        ws["A3"].alignment = top_align
        if domains:
            for i, domain in enumerate(domains):
                ws[f"B{3 + i}"] = domain
                ws[f"B{3 + i}"].font = regular_font
            ws.merge_cells(f"A3:A{3 + len(domains) - 1}")

        # Sending Ips
        ip_start = 3 + len(domains)
        ws[f"A{ip_start}"] = "Sending IP Addresses"
        ws[f"A{ip_start}"].alignment = top_align
        ws[f"A{ip_start}"].font = header_font
        if ips:
            for i, ip in enumerate(ips):
                ws[f"B{ip_start + i}"] = ip
                ws[f"B{ip_start + i}"].font = regular_font
            ws.merge_cells(f"A{ip_start}:A{ip_start + len(ips) - 1}")

        # Templates
        template_start = 3 + len(domains) + (1 if not ips else len(ips)) + 1
        ws[f"A{template_start}"] = "Template Subject"
        ws[f"B{template_start}"] = "Deception Level"
        ws[f"A{template_start}"].font = header_font
        ws[f"B{template_start}"].font = header_font
        for i, template in enumerate(templates):
            ws[f"A{template_start + i + 1}"] = template["subject"]
            ws[f"A{template_start + i + 1}"].font = regular_font
            ws[f"B{template_start + i + 1}"] = get_deception_level(
                template["deception_score"]
            ).title()
            ws[f"B{template_start + i + 1}"].font = regular_font

        # Expand columns in workbook
        for cells in ws.columns:
            new_column_length = max(len(str(cell.value)) for cell in cells)
            new_column_letter = get_column_letter(cells[0].column)
            if new_column_length > 0:
                ws.column_dimensions[new_column_letter].width = new_column_length * 1.5

        filepath = f"/var/www/tmp_safelist_{subscription_id}.xslx"
        wb.save(filepath)

        # Protect excel workbook
        if reporting_password:
            args = [
                "/var/www/msoffice-crypt.exe",
                "-e",
                "-p",
                reporting_password,
                filepath,
                filepath,
            ]
            subprocess.run(args)  # nosec

        try:
            return send_file(
                filepath,
                as_attachment=True,
                attachment_filename="safelist_export.xlsx",
            )
        except Exception as e:
            logging.exception(e)
            raise e
        finally:
            logging.info(f"Deleting file {filepath}")
            os.remove(filepath)
